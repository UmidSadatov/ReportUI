from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pyexcel as p
import divisions
import db_manage as db
import json


def convert_to_xlsx(filename: str):
    if filename.split('.')[-1] == 'xls':
        p.save_book_as(file_name=filename,
                       dest_file_name=filename + 'x')
        return filename + 'x'




class ExcelBook:
    """
    Класс для существующего Эксель-файла

    Поля:
    filename: имя файла (указывается в конструкторе)
    book - объект Workbook (openpyexcel)
    number_of_sheets - количество листов
    """    

    # Конструктор для открытия указанного файла
    def __init__(self, filename: str):
        extension = filename.split('.')[-1].lower()

        if extension not in ['xlsx', 'xls']:
            raise ValueError('Файл не выбран!')
        elif extension == 'xls':
            self.filename = convert_to_xlsx(filename)
        else:
            self.filename = filename
        self.book = load_workbook(self.filename, data_only=True)
        self.number_of_sheets = len(self.book.worksheets)

    def get_cell_value(self, cell: str, sheet_index=0, round_value=True):

        try:
            sheet = self.book.worksheets[sheet_index]
        except IndexError:
            return None

        value = sheet[cell].value if sheet[cell].value is not None else 0

        if type(value) is float and round_value:
            value = round(value, 2)

        return value

    def set_cell_value(
            self, cell, value, sheet_index=0, bold=False, 
            center=False, rotate=False, fill=None
    ):
        sheet = self.book.worksheets[sheet_index]
        sheet[cell].value = value
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        sheet[cell].border = thin_border
        if bold:
            sheet[cell].font = Font(bold=True)
        
        sheet[cell].alignment = Alignment(
            horizontal='center' if center else None, 
            vertical='center' if center else None,
            text_rotation=90 if rotate else None
        )

        if fill is not None:
            sheet[cell].fill = fill

        # self.book.save(self.filename)

    def get_max_row(self, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        return sheet.max_row
    
    def get_max_col_int(self, sheet_index=0):
        sheet = self.book.worksheets[sheet_index]
        return sheet.max_column
    
    def rename_sheet_by_index(self, sheet_index:int, title:str):
        self.book.worksheets[sheet_index].title = title
        # self.book.save(self.filename)
    
    def add_sheet(self, title:str):
        self.book.create_sheet(title)
        self.number_of_sheets += 1
        # self.book.save(self.filename)



class DistributorsFilePrimary:
    def __init__(
        self, 
        filename:str, 
        distributor: str,
        sheet_index = 0,
        name_column = None, 
        balance_beginning_column = None, 
        incomes_column = None, 
        sold_number_column = None, 
        balance_end_column = None, 
        start_row_number = 1
    ):
        self.filename = filename
        self.book = ExcelBook(filename)
        self.distributor = distributor        
        self.sheet_index = sheet_index
        self.name_column = name_column
        self.balance_beginning_column = balance_beginning_column
        self.incomes_column = incomes_column
        self.sold_number_column = sold_number_column
        self.balance_end_column = balance_end_column
        self.start_row_number = start_row_number
    
    def render(self):
        not_found_names = set()
        result_dict = {
            "Total": {
                "balance_beginning": 0,
                "incomes": 0,
                "sold_number": 0,
                "balance_end": 0
            }
        }

        for row in range(self.start_row_number, self.book.get_max_row(sheet_index=self.sheet_index)):
            
            unique_name = self.book.get_cell_value(f"{self.name_column}{row}")
            general_name = db.get_general_name(str(unique_name))
            if general_name is None:
                not_found_names.add(unique_name)
                continue

            balance_beginning = int(self.book.get_cell_value(f"{self.balance_beginning_column}{row}", sheet_index=self.sheet_index))

            incomes = int(self.book.get_cell_value(f"{self.incomes_column}{row}", sheet_index=self.sheet_index))
            
            balance_end = int(self.book.get_cell_value(f"{self.balance_end_column}{row}", sheet_index=self.sheet_index))

            sold_number = balance_beginning + incomes - balance_end

            if general_name in result_dict:
                result_dict[general_name]["balance_beginning"] += balance_beginning
                result_dict[general_name]["incomes"] += incomes
                result_dict[general_name]["sold_number"] += sold_number
                result_dict[general_name]["balance_end"] += balance_end
            else:
                result_dict[general_name] = {
                    "balance_beginning": balance_beginning,
                    "incomes": incomes,
                    "sold_number": sold_number,
                    "balance_end": balance_end
                }
            
            result_dict["Total"]["balance_beginning"] += balance_beginning
            result_dict["Total"]["incomes"] += incomes
            result_dict["Total"]["sold_number"] += sold_number
            result_dict["Total"]["balance_end"] += balance_end

        return result_dict, not_found_names


class DistributorsFileSecondary:
    def __init__(
            self, 
            filename:str, 
            distributor: str,
            sheet_index = 0,
            name_column = None, 
            region1_column = None, 
            region2_column = None, 
            sold_number_column = None, 
            client_column = None, 
            start_row_number = 1
    ):
        self.filename = filename
        self.book = ExcelBook(filename)
        self.distributor = distributor

        if distributor == "Не выбрано":
            raise ValueError("Дистрибьютор не выбран!")
        
        if self.distributor in ("Akmal Pharm", "Tabletka", "Pharma Cosmos"):
            self.rendering_mode = self.distributor
        else:
            self.rendering_mode = "Usual"

        self.sheet_index = sheet_index

        self.name_column = name_column
        self.region1_column = region1_column
        self.region2_column = region2_column

        if self.distributor == "Akmal Pharm":
            self.sold_number_column = 'P'
        else:
            self.sold_number_column = sold_number_column

        self.client_column = client_column
        self.start_row_number = start_row_number
    

    def render(self):

        result_dict = db.get_all_names()

        not_found_names = set()
        not_found_regions = set()

        if self.rendering_mode == "Usual":

            if self.name_column is None:
                raise ValueError("Введите столбец наименований!")
            elif self.region1_column is None:
                raise ValueError("Введите столбец регионов 1!")
            elif self.sold_number_column is None:
                raise ValueError("Введите столбец продаж!")
            elif self.start_row_number is None:
                raise ValueError("Введите с какой строки начать считывать!")

            for row in range(self.start_row_number, self.book.get_max_row(self.sheet_index)+1):
            
                unique_name = self.book.get_cell_value(f"{self.name_column}{row}", sheet_index=self.sheet_index)
                general_name = db.get_general_name(str(unique_name))
                
                if general_name is None:
                    not_found_names.add(unique_name)
                    continue

                sold_count = self.book.get_cell_value(f"{self.sold_number_column}{row}", sheet_index=self.sheet_index)

                if sold_count < 0:
                    sold_count = -sold_count

                reg1 = str(self.book.get_cell_value(f"{self.region1_column}{row}", sheet_index=self.sheet_index))
                reg2 = str(self.book.get_cell_value(f"{self.region2_column}{row}", sheet_index=self.sheet_index)) if self.region2_column is not None else ''
                
                general_region = db.get_general_region(reg1)
                if general_region is None:
                    general_region = db.get_general_region(reg2) if len(reg2) else None
                if general_region is None:
                    general_region = db.get_general_region(reg1 + ' ' + reg2)
                if general_region is None:
                    not_found_regions.add(reg1)
                    if len(reg2):
                        not_found_regions.add(reg2)

                    group = db.get_group_of_gen_name(general_name)
                    result_dict[group][general_name]['Total'] += sold_count
                    result_dict[group][general_name]['Неизвестный регион'] += sold_count
                    continue
                
                if self.client_column is not None:
                    client = self.book.get_cell_value(f"{self.client_column}{row}", sheet_index=self.sheet_index)

                    division = divisions.divide(client, sold_number=sold_count)

                    if division is not None:
                        div_regs = division[0]
                        div_chunk = division[1]
                        
                        for reg in div_regs:
                            group = db.get_group_of_gen_name(general_name)
                            result_dict[group][general_name][reg] += div_chunk
                            result_dict[group][general_name]['Total'] += div_chunk

                    else:
                        group = db.get_group_of_gen_name(general_name)
                        result_dict[group][general_name][general_region] += sold_count
                        result_dict[group][general_name]['Total'] += sold_count
                
                else:
                    group = db.get_group_of_gen_name(general_name)
                    result_dict[group][general_name][general_region] += sold_count
                    result_dict[group][general_name]['Total'] += sold_count

        elif self.rendering_mode == "Akmal Pharm":
            general_region = None
            for row in range(self.start_row_number, self.book.get_max_row(self.sheet_index)+1):
                if self.book.get_cell_value(f"B{row}") == 0:
                    general_region = db.get_general_region(self.book.get_cell_value(f"A{row}"))
                    if general_region is None:
                        not_found_regions.add(general_region)
                    continue

                else:
                    name = self.book.get_cell_value(f"B{row}")
                    general_name = db.get_general_name(name)
                    sold_count = self.book.get_cell_value(f"{self.sold_number_column}{row}")                    

                    if general_name is None:
                        not_found_names.add(name)
                        continue

                    if general_region is None:
                        result_dict[group][general_name]['Неизвестный регион'] += sold_count
                        result_dict[group][general_name]['Total'] += sold_count
                        continue

                    else:
                        group = db.get_group_of_gen_name(general_name)                        
                        result_dict[group][general_name][general_region] += sold_count
                        result_dict[group][general_name]['Total'] += sold_count

        # elif self.rendering_mode == "Tabletka":
        #     columns_and_regions = {}
        #     ignored_column = None

        #     for row in range(1, self.book.get_max_row(self.sheet_index)):
        #         if db.get_general_region(self.book.get_cell_value(f"B{row}", sheet_index=self.sheet_index)) is not None:
        #             columns_and_regions = {}
        #             for col in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        #                 reg = self.book.get_cell_value(f"{col}{row}", sheet_index=self.sheet_index)
        #                 general_region = db.get_general_region(reg)
        #                 if general_region is not None:
        #                     columns_and_regions[col] = general_region
        #                 else:
        #                     not_found_regions.add(reg)
        #                     if "Общий итог" in str(reg):
        #                         ignored_column = col
                
        #         elif self.book.get_cell_value(f"A{row}", sheet_index=self.sheet_index) != 0:
        #             name = str(self.book.get_cell_value(f"A{row}", sheet_index=self.sheet_index))

        #             while len(name) and name[-1] == " ": name = name[:-1]
        #             while len(name) and name[0] == " ": name = name[1:]

        #             general_name = db.get_general_name(name)
        #             if general_name is not None:
        #                 # sold_count_row = row + 3
        #                 # while self.book.get_cell_value(f"A{sold_count_row}", sheet_index=self.sheet_index) != "Сумма по полю расход":
        #                 #     sold_count_row += 1
        #                 group = db.get_group_of_gen_name(general_name)
        #                 for col in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        #                     if col == ignored_column:
        #                         continue
        #                     sold_count = self.book.get_cell_value(f"{col}{row + 3}", sheet_index=self.sheet_index)
        #                     if col in columns_and_regions.keys():
        #                         genreg = columns_and_regions[col]                                
        #                         result_dict[group][general_name][genreg] += sold_count                                
        #                     else:
        #                         result_dict[group][general_name]["Неизвестный регион"] += sold_count
        #                     result_dict[group][general_name]["Total"] += sold_count
        #             else:
        #                 not_found_names.add(name)
        

        elif self.rendering_mode == "Tabletka":

            if self.sold_number_column is None:
                raise ValueError("Введите столбец продаж!")
            elif self.start_row_number is None:
                raise ValueError("Введите с какой строки начать считывать!")
            
            sold_count_row_shift = int(self.sold_number_column.replace("+", ""))
            row = self.start_row_number
            max_row = self.book.get_max_row(sheet_index=self.sheet_index)
            max_col_int = self.book.get_max_col_int(sheet_index=self.sheet_index)
            
            while row <= max_row:
                guess_name = self.book.get_cell_value(f"A{row}", sheet_index=self.sheet_index)
                general_name = db.get_general_name(guess_name)
                skipping_cols = set()

                if general_name is not None:
                    row += sold_count_row_shift
                    for region_col_int in range(2, max_col_int):

                        region_col = get_column_letter(region_col_int)                        
                        group = db.get_group_of_gen_name(general_name)
                        
                        try:
                            sold_count = int(self.book.get_cell_value(f"{region_col}{row}", sheet_index=self.sheet_index))
                        except ValueError:
                            continue

                        if region_col in skipping_cols:
                            result_dict[group][general_name]["Неизвестный регион"] += sold_count
                            result_dict[group][general_name]["Total"] += sold_count
                            continue

                        guess_region = \
                            self.book.get_cell_value(f"{region_col}{self.start_row_number-1}", sheet_index=self.sheet_index)
                        
                        general_region = db.get_general_region(guess_region)                                                

                        if general_region is not None:                            
                            result_dict[group][general_name][general_region] += sold_count
                        else:
                            skipping_cols.add(region_col)
                            result_dict[group][general_name]["Неизвестный регион"] += sold_count
                            
                        result_dict[group][general_name]["Total"] += sold_count

                else:
                    not_found_names.add(guess_name)
                
                row += 1







        elif self.rendering_mode == "Pharma Cosmos":
            
            with open("json_files/ignored_suppliers.json", 'r', encoding='utf-8') as file:
                ignored_suppliers = json.load(file)
                
            reg_cols = "DFGHIJKLMNOPQRSTUVWXYZ"

            columns_and_regions = {}
            not_found_regs_cols = []

            for col in reg_cols:
                if self.book.get_cell_value(f"{col}5") == "Итого":
                    break
                else:
                    reg = self.book.get_cell_value(f"{col}5")
                    gen_reg = db.get_general_region(reg)
                    if gen_reg is not None:
                        columns_and_regions[col] = gen_reg
                    else:
                        not_found_regions.add(reg)
                        not_found_regs_cols.append(col)
            
            accept_data = False

            for row in range(7, self.book.get_max_row()+1):
                a_col_value = self.book.get_cell_value(f"A{row}")

                if a_col_value == a_col_value.upper():
                    if a_col_value in ignored_suppliers:
                        accept_data = False
                    else:
                        accept_data = True
                
                else:
                    if accept_data:
                        general_name = db.get_general_name(a_col_value)
                        if general_name is not None:
                            group = db.get_group_of_gen_name(general_name)

                            for col in columns_and_regions.keys():
                                gen_reg = columns_and_regions[col]
                                sold_count = self.book.get_cell_value(f"{col}{row}")
                                result_dict[group][general_name][gen_reg] += sold_count
                                result_dict[group][general_name]["Total"] += sold_count
                            
                            for nfr_col in not_found_regs_cols:
                                sold_count = self.book.get_cell_value(f"{nfr_col}{row}")
                                result_dict[group][general_name]["Неизвестный регион"] += sold_count
                                result_dict[group][general_name]["Total"] += sold_count
                                
                        else:
                            not_found_names.add(a_col_value)
                            continue
                    else:
                        continue
        
        # result_dict[group][general_name][region | 'Total' | 'Неизвестный регион']

        with open("json_files/employees.json", 'r', encoding='utf-8') as file:
            employees = json.load(file)
        
        return result_dict, not_found_names, not_found_regions




class TotalReportPrimary:
    def __init__(self, *dist_files: DistributorsFilePrimary):
        self.dist_files = dist_files
        self.result_dict = {
            "Total": db.get_all_names(include_regions=False)
        }
        self.not_found_names = set()
    
    def render(self):
        not_found_names = set()
        total_result_dict = dict()
        """
        {
            "Total": {...},
            "Pharm_Luxe":{
                            "group1": {
                                        "product1": {
                                                        "balance_beginning":0,
                                                        "incomes":0,
                                                        "sold_count":0,
                                                        "balance_end":0,
                                                    }
                                    },
                            "group2": {
                                        "product1": {
                                                        "balance_beginning":0,
                                                        "incomes":0,
                                                        "sold_count":0,
                                                        "balance_end":0,
                                                    }
                                    },
                            ...,

                            "Total": {
                                        "balance_beginning":0,
                                        "incomes":0,
                                        "sold_count":0,
                                        "balance_end":0,
                                    }
                        },
            ...
        }
        """

        for dist_file in self.dist_files:
            dist_result_dict, dist_not_found_names = dist_file.render()

            if dist_file.distributor not in total_result_dict:
                total_result_dict[dist_file.distributor] = db.get_all_names(include_regions=False)
            
            for name in dist_result_dict.keys():
                if name != "Total":
                    group = db.get_group_of_gen_name(name)
                    total_result_dict[group][name]
            

            



            for name in dist_not_found_names:
                not_found_names.add(name)
            


            
                




                        
class TotalReportSecondary:
    def __init__(self, *dist_files: DistributorsFileSecondary):
        self.dist_files = dist_files
        self.result_dict = {
            "Total": db.get_all_names()
        }
        self.not_found_names = set()
        self.not_found_regions = set()
    
    def render(self):
        for dist_file in self.dist_files:
            dist_result_dict, dist_not_found_names, dist_not_found_regions = dist_file.render()

            self.not_found_names.update(dist_not_found_names)
            self.not_found_regions.update(dist_not_found_regions)

            for group in dist_result_dict.keys():
                for name in dist_result_dict[group].keys():
                    for region in dist_result_dict[group][name].keys():
                        sold_count = dist_result_dict[group][name][region]
                        self.result_dict["Total"][group][name][region] += sold_count

            if dist_file.distributor not in self.result_dict.keys():
                self.result_dict[dist_file.distributor] = dist_result_dict
            else:
                for group in dist_result_dict.keys():
                    for name in dist_result_dict[group].keys():
                        for region in dist_result_dict[group][name].keys():
                            sold_count = dist_result_dict[group][name][region]
                            self.result_dict[dist_file.distributor][group][name][region] += sold_count
        
        return self.result_dict, self.not_found_names, self.not_found_regions
    
    def make_excel(self, new_file_name):

        self.render()

        # now = datetime.now()
        # new_file_name = f"Reports/Report-{now.strftime("%d%m%Y%H%M%S%f")[:-3]}.xlsx"

        wb = Workbook()
        wb.save(new_file_name)

        result_book = ExcelBook(new_file_name)

        green_fill = PatternFill(
            start_color='D5FFDC',
            end_color='D5FFDC',
            fill_type='solid'
        )

        yellow_fill = PatternFill(
            start_color='F8FFD7',
            end_color='F8FFD7',
            fill_type='solid'
        )

        result_book.rename_sheet_by_index(0, "Total")

        for ditributor_name in list(self.result_dict.keys())[1:]:
            result_book.add_sheet(ditributor_name)
        
        for dist_name in self.result_dict:
            sheet = result_book.book[dist_name]
            sheet_index = list(self.result_dict).index(dist_name)
            
            # Заголовки:

            sheet.column_dimensions['A'].width = 10
            result_book.set_cell_value('A1', "Группа", sheet_index=sheet_index, bold=True, center=True)

            sheet.column_dimensions['B'].width = 40
            result_book.set_cell_value('B1', "Наименование", sheet_index=sheet_index, bold=True, center=True)

            region_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']

            regions = db.get_all_regions()

            regions.insert(0, "Total")
            regions.append("Неизвестный регион")

            for reg_index in range(0, len(regions)):
                region = regions[reg_index]
                column = region_columns[reg_index]
                sheet[f"{column}1"].alignment = Alignment(horizontal="center")
                result_book.set_cell_value(
                    f"{column}1", 
                    region, 
                    sheet_index=sheet_index, 
                    bold=True,
                    rotate=True
                )
            
            # Данные:
            
            curr_row = 2
            current_fill = None
            for group in list(self.result_dict[dist_name].keys()):

                group_index = list(self.result_dict[dist_name].keys()).index(group)

                if group_index % 2:
                    current_fill = green_fill
                else:
                    current_fill = yellow_fill

                for name in list(self.result_dict[dist_name][group].keys()):
                    result_book.set_cell_value(f"A{curr_row}", group, sheet_index=sheet_index, fill=current_fill, center=True)
                    result_book.set_cell_value(f"B{curr_row}", name, sheet_index=sheet_index, fill=current_fill)
                    
                    for reg_index in range(0, len(regions)):
                        region = regions[reg_index]
                        column = region_columns[reg_index]
                        data = int(self.result_dict[dist_name][group][name][region])
                        result_book.set_cell_value(f"{column}{curr_row}", data, sheet_index=sheet_index)
                    curr_row += 1

        result_book.book.save(new_file_name)

        # with open('newdata.json', 'w', encoding='utf-8') as file:
        #     json.dump(self.result_dict, file, ensure_ascii=False, indent=4)

        return self.not_found_names, self.not_found_regions


                   






                




        

        
